#! python3
# cellInverter.py
# A program that will invert the row and column of the cells of the given spreadsheet.

import openpyxl, sys, os

#function for when program is run in the command line
def execute_standalone():
    global filename
    
    #If a direct path is not used for the filename, then it will default to the path that the command prompt
    #opens to
    filename = sys.argv[1]

#main

#If program is run in IDLE shell
if 'idlelib.run' in sys.modules:
    print('Enter name of excel file') 
    filename = input()  
    
#Otherwise assume program is run from command line
else:
    execute_standalone()

print('Opening workbook ... ')
wb = openpyxl.load_workbook(filename) #open excel spreadsheet
sheet = wb.active #store instance of the active sheet
fileData = {}

print('Reading rows ...')
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        #store all the data in the sheet to fileData
        fileData.setdefault(row, {})
        fileData[row].setdefault(column)
        fileData[row][column] = sheet.cell(row=row, column=column).value

blankWb = openpyxl.Workbook() #open a blank workbook
blankSheet = blankWb.active


for row in range(1, sheet.max_column + 1):
    for column in range(1, sheet.max_row + 1):
        #Do the actual inverting here, swapping the data in rows with those in the columns
        blankSheet.cell(row=row, column=column).value = fileData[column][row]

blankWb.save('cellsInverted' + filename) #save to a new excel spreadsheet

