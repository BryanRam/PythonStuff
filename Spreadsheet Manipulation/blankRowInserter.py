#! python3
#blankRowInserter.py
#A program that inserts M blank rows starting at row N

import openpyxl, sys, os

#function for when program is run in the command line
def execute_standalone():
    global start, numRows, filename
    start = int(sys.argv[1])
    numRows = int(sys.argv[2])
    #If a direct path is not used for the filename, then it will default to the path that the command prompt
    #opens to
    filename = sys.argv[3]

#main

#If program is run in IDLE shell
if 'idlelib.run' in sys.modules:
    print('Enter name of excel file') 
    filename = input()
    print('Enter num of the start row')
    start = int(input())
    print('Enter number of rows')
    numRows = int(input())
   
    
#Otherwise assume program is run from command line
else:
    execute_standalone()

print('Opening workbook ... ')
print(os.getcwd())
wb = openpyxl.load_workbook(filename)
sheet = wb.active
fileData = {}

print('Reading rows ...')
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        fileData.setdefault(row, {})
        fileData[row].setdefault(column)
        fileData[row][column] = sheet.cell(row=row, column=column).value

blankWb = openpyxl.Workbook()#Open a blank workbook
blankSheet = blankWb.active


for row in range(1, sheet.max_row + 1):
    if row < start:
        #output sheet data onto the new workbook as normal, until you reach the start of the blank rows
        for column in range(1, sheet.max_column + 1):
            blankSheet.cell(row=row, column=column).value = fileData[row][column]
    else:
        #when that happens, skip 'numRows' rows then output the rest of the sheet data
        for column in range(1, sheet.max_column + 1):
            blankSheet.cell(row=row+numRows, column=column).value = fileData[row][column]

blankWb.save('BlankRowInserted' + filename) #save to a new excel spreadsheet
