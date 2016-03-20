#python3
#multiplicationTable.py
#This program takes a number N from the command line or as prompted
#and creates an N×N multiplication table in an Excel spreadsheet.

import openpyxl, sys
from openpyxl.styles import Font, Style, Side, Border

#function for when program is run in the command line
def execute_standalone():
    global num
    num = sys.argv[1]
    

#main

#If program is run in IDLE shell
if 'idlelib.run' in sys.modules:
    print('Enter a number')
    num = int(input())
    
#Otherwise assume program is run from command line
else:
    execute_standalone()


wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, num + 1):
    #Create the labels, which should be in bold
    sheet.cell(row=1, column=(i+1)).font = Font(bold=True)
    sheet.cell(row=1,column=(i+1)).value = i
    sheet.cell(row=(i+1), column=1).font = Font(bold=True)
    sheet.cell(row=(i+1),column=1).value = i

    for j in range(1, num + 1):
        #Populate the spreadsheet with the products
        sheet.cell(row=i+1, column=j+1).value = i * j
        if(i == num):
            #insert a border along the bottom of cells when reached
            sheet.cell(row=i+1, column=(j+1)).border = Border(left=Side(style='thin'),
                                                              right=Side(style='thin'),
                                                              bottom=Side(style='thin'))
        else:
            #otherwise insert borders along the left and right of each cell
            sheet.cell(row=i+1, column=(j+1)).border = Border(left=Side(style='thin'),
                                                              right=Side(style='thin'))
        
#freeze the top row and leftmost column, which contain the labels                                                              
sheet.freeze_panes = 'B2'


wb.save('FreezeTables.xlsx')        
